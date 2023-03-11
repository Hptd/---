import openpyxl
import pandas
import os
import re


class DataCleansing(object):
    def __init__(self):
        """
        所有的代码都可以不用看，直接把每个省的文件夹放在和.py
        文件同一目录下，放哪个省，就直接替换掉下面 “贵州省”
        其余的不需要更改。
        """
        self.csv_path = "./四川省/"
        self.files = os.listdir(self.csv_path)

    def filter_data(self, csv_name):
        csv_file = pandas.read_csv(self.csv_path + csv_name)
        row = 0
        wb_row = 2
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet["A1"] = "名称"
        sheet["B1"] = "大类"
        sheet["C1"] = "中类"
        sheet["D1"] = "经度"
        sheet["E1"] = "纬度"
        sheet["F1"] = "省份"
        sheet["G1"] = "城市"
        sheet["H1"] = "区域"
        while True:
            try:
                value_list = []
                value = csv_file.at[row, '大类']
                if value == "旅游景点":
                    value_list.append(csv_file.at[row, "名称"])
                    value_list.append(csv_file.at[row, "大类"])
                    value_list.append(csv_file.at[row, "中类"])
                    value_list.append(csv_file.at[row, "经度"])
                    value_list.append(csv_file.at[row, "纬度"])
                    value_list.append(csv_file.at[row, "省份"])
                    value_list.append(csv_file.at[row, "城市"])
                    value_list.append(csv_file.at[row, "区域"])

                    list_idx = 0
                    for column in "ABCDEFGH":
                        wb_cell_name = column + str(wb_row)
                        wb_cell = sheet[wb_cell_name]
                        wb_cell.value = value_list[list_idx]
                        list_idx += 1
                    wb_row += 1
                row += 1

            except KeyError:
                wb.save(csv_name[:-4] + ".xlsx")
                break

    def loop_read_csv(self):
        for file in self.files:
            print(file)
            self.filter_data(file)

    def merge_excel(self):
        self.loop_read_csv()

        all_files = os.listdir(".")
        xlsx_files = [file for file in all_files if file.endswith(".xlsx")]
        all_wb = openpyxl.Workbook()
        all_wb_sheet = all_wb.active
        all_wb_sheet["A1"] = "名称"
        all_wb_sheet["B1"] = "大类"
        all_wb_sheet["C1"] = "中类"
        all_wb_sheet["D1"] = "经度"
        all_wb_sheet["E1"] = "纬度"
        all_wb_sheet["F1"] = "省份"
        all_wb_sheet["G1"] = "城市"
        all_wb_sheet["H1"] = "区域"
        for file in xlsx_files:
            xlsx_wb = openpyxl.load_workbook(file)
            xlsx_wb_sheet = xlsx_wb.active
            for row in xlsx_wb_sheet.iter_rows(min_row=2):
                all_wb_sheet.append([cell.value for cell in row])
        all_xlsx_name = re.sub(r"\./|/", "", self.csv_path) + ".xlsx"
        all_wb.save(all_xlsx_name)
        print("Done.")


if __name__ == '__main__':
    DataCleansing().merge_excel()
